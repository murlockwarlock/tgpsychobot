import io
import csv
import re
import zipfile
from xml.etree import ElementTree
try:
    from docx import Document
except ModuleNotFoundError:
    Document = None
try:
    from pypdf import PdfReader
except ModuleNotFoundError:
    PdfReader = None
try:
    import openpyxl
except ModuleNotFoundError:
    openpyxl = None

from universal_tests import HORIZONTAL_VALUES, is_truthy, json_dumps, normalize_variable_name


async def parse_file(file_bytes: io.BytesIO, filename: str) -> str:
    import asyncio
    text = ""
    file_extension = filename.split('.')[-1].lower()

    if file_extension == 'txt' or file_extension == 'md':
        text = file_bytes.read().decode('utf-8')
    elif file_extension == 'pdf':
        if PdfReader is None:
            raise RuntimeError("Для чтения PDF нужен пакет pypdf.")
        def parse_pdf():
            nonlocal text
            reader = PdfReader(file_bytes)
            for page in reader.pages:
                text += page.extract_text() or ""
        await asyncio.to_thread(parse_pdf)
    elif file_extension == 'docx':
        if Document is None:
            raise RuntimeError("Для чтения DOCX нужен пакет python-docx.")
        def parse_docx():
            nonlocal text
            doc = Document(file_bytes)
            for para in doc.paragraphs:
                text += para.text + '\n'
        await asyncio.to_thread(parse_docx)
    elif file_extension == 'xlsx':
        if openpyxl is None:
            raise RuntimeError("Для чтения XLSX нужен пакет openpyxl.")
        def parse_xlsx():
            nonlocal text
            workbook = openpyxl.load_workbook(file_bytes)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text += str(cell.value) + ' '
                    text += '\n'
        await asyncio.to_thread(parse_xlsx)

    return text.strip()


async def parse_questions_file(file_bytes: io.BytesIO, filename: str) -> dict[str, list[dict]]:
    import asyncio
    questions = []
    formulas = []
    file_extension = filename.split('.')[-1].lower()

    if file_extension == 'xlsx':
        def parse_xlsx_questions():
            if openpyxl is not None:
                workbook = openpyxl.load_workbook(file_bytes, data_only=False)
                sheets = [(ws.title, list(ws.iter_rows(min_row=1, values_only=True))) for ws in workbook.worksheets]
            else:
                sheets = _read_xlsx_sheets(file_bytes.read())
            rows = next((sheet_rows for title, sheet_rows in sheets if title.strip().lower() not in {"formulas", "формулы"}), [])
            question_rows, inline_formula_rows = _partition_question_and_formula_rows(rows)
            questions.extend(_parse_question_rows(question_rows))
            formulas.extend(_parse_formula_rows(inline_formula_rows))
            formula_rows = next((sheet_rows for title, sheet_rows in sheets if title.strip().lower() in {"formulas", "формулы"}), None)
            if formula_rows:
                formulas.extend(_parse_formula_rows(formula_rows))

        await asyncio.to_thread(parse_xlsx_questions)

    elif file_extension in {'txt', 'csv'}:
        content = _decode_tabular_text(file_bytes.read())
        rows = _read_delimited_rows(content)
        question_rows, formula_rows = _partition_question_and_formula_rows(rows)
        questions.extend(_parse_question_rows(question_rows))
        formulas.extend(_parse_formula_rows(formula_rows))

    return {"questions": questions, "formulas": formulas}


async def parse_formulas_file(file_bytes: io.BytesIO, filename: str) -> list[dict]:
    import asyncio

    extension = filename.rsplit('.', 1)[-1].lower()
    if extension == 'xlsx':
        def parse_xlsx_formulas():
            if openpyxl is not None:
                workbook = openpyxl.load_workbook(file_bytes, data_only=False)
                sheet = next(
                    (ws for ws in workbook.worksheets if ws.title.strip().lower() in {"formulas", "формулы"}),
                    workbook.active,
                )
                rows = list(sheet.iter_rows(min_row=1, values_only=True))
            else:
                sheets = _read_xlsx_sheets(file_bytes.read())
                rows = next((sheet_rows for title, sheet_rows in sheets if title.strip().lower() in {"formulas", "формулы"}), sheets[0][1] if sheets else [])
            _, formula_rows = _split_formula_section(rows)
            return _parse_formula_rows(formula_rows or rows)

        return await asyncio.to_thread(parse_xlsx_formulas)

    if extension in {'txt', 'csv'}:
        content = _decode_tabular_text(file_bytes.read())
        rows = _read_delimited_rows(content)
        _, formula_rows = _split_formula_section(rows)
        return _parse_formula_rows(formula_rows or rows)

    raise ValueError("Поддерживаются только файлы .xlsx, .csv и .txt.")


def _parse_question_rows(rows) -> list[dict]:
    rows = [tuple(row or []) for row in rows]
    first_real_index = next((index for index, row in enumerate(rows) if any(_cell_text(cell) for cell in row)), None)
    first_real = rows[first_real_index] if first_real_index is not None else ()
    new_format = _looks_like_new_test_format(first_real)
    has_header = _looks_like_question_header(first_real)
    schema = _build_question_schema(first_real) if has_header else None
    result = []
    for row_index, row in enumerate(rows):
        if not any(_cell_text(cell) for cell in row):
            continue
        if has_header and row_index == first_real_index:
            continue
        if new_format:
            parsed = _parse_new_question_row(row, len(result), schema)
        else:
            parsed = _parse_legacy_question_row(row)
        if parsed:
            result.append(parsed)
    return result


def _split_formula_section(rows) -> tuple[list, list]:
    question_rows = []
    formula_rows = []
    target = question_rows
    for row in rows:
        first = str(row[0]).strip().lower() if row else ""
        if first in {"[formulas]", "formulas", "формулы", "[формулы]"}:
            target = formula_rows
            continue
        target.append(row)
    return question_rows, formula_rows


def _partition_question_and_formula_rows(rows) -> tuple[list, list]:
    first_real = next((row for row in rows if row and any(_cell_text(cell) for cell in row)), ())
    if _looks_like_formula_header(first_real):
        return [], list(rows)
    return _split_formula_section(rows)


def _looks_like_formula_header(row) -> bool:
    first = _cell_text(row[0] if len(row) > 0 else None).lower()
    second = _cell_text(row[1] if len(row) > 1 else None).lower()
    return first in {"переменная", "variable", "name", "название"} and second in {"формула", "formula", "expression"}


def _parse_legacy_question_row(row) -> dict | None:
    if not _cell_text(row[0] if len(row) > 0 else None):
        return None
    if _cell_text(row[0]).lower().startswith('вопрос') or _cell_text(row[1] if len(row) > 1 else None).lower().startswith('категория'):
        return None
    return {
        'text': _cell_text(row[0]),
        'category': _cell_text(row[1] if len(row) > 1 else None) or 'general',
        'is_reverse': is_truthy(row[2] if len(row) > 2 else None),
    }


def _parse_new_question_row(row, index: int, schema: dict | None = None) -> dict | None:
    question_text = _cell_text(row[1] if len(row) > 1 else None)
    if not question_text:
        return None
    options = _parse_answer_options_from_row(row, schema) if schema else _parse_answer_options(row[6:])
    return {
        'text': question_text,
        'category': 'general',
        'is_reverse': False,
        'comment': _cell_text(row[2] if len(row) > 2 else None) or None,
        'variable_name': normalize_variable_name(row[3] if len(row) > 3 else None, f"answer_{index + 1:02d}"),
        'allow_custom_answer': is_truthy(row[4] if len(row) > 4 else None),
        'buttons_layout': 'horizontal' if _cell_text(row[5] if len(row) > 5 else None).lower() in HORIZONTAL_VALUES else 'vertical',
        'answer_options_json': json_dumps(options) if options else None,
    }


def _parse_answer_options(cells) -> list[dict]:
    values = [_cell_text(cell) for cell in cells if _cell_text(cell)]
    if not values:
        return []

    numbers = [_to_float_or_none(value) for value in values]
    if all(item is not None for item in numbers):
        return [{"text": value, "value": number} for value, number in zip(values, numbers)]

    split_at = None
    for candidate in range(1, len(values)):
        labels = values[:candidate]
        numeric_tail = values[candidate:]
        if len(numeric_tail) == len(labels) and all(_to_float_or_none(item) is not None for item in numeric_tail):
            split_at = candidate
            break

    if split_at is None:
        return [{"text": value, "value": _to_float_or_none(value)} for value in values]

    labels = values[:split_at]
    numeric_tail = values[split_at:]
    return [{"text": label, "value": _to_float_or_none(numeric_tail[index])} for index, label in enumerate(labels)]


def _parse_answer_options_from_row(row, schema: dict) -> list[dict]:
    option_columns = schema.get("option_columns", [])
    value_columns = schema.get("value_columns", [])
    button_columns = schema.get("button_columns", [])
    options = []
    for option_index, column in enumerate(option_columns):
        label = _cell_text(row[column] if column < len(row) else None)
        if not label:
            continue
        value_column = value_columns[option_index] if option_index < len(value_columns) else None
        value_text = _cell_text(row[value_column] if value_column is not None and value_column < len(row) else None)
        option = {"text": label, "value": _to_float_or_none(value_text or label)}
        button_column = button_columns[option_index] if option_index < len(button_columns) else None
        button_text = _cell_text(row[button_column] if button_column is not None and button_column < len(row) else None)
        if button_text:
            option["button_text"] = button_text
        options.append(option)
    return options


def _parse_formula_rows(rows) -> list[dict]:
    formulas = []
    for row in rows:
        if not row or not any(_cell_text(cell) for cell in row):
            continue
        first = _cell_text(row[0])
        second = _cell_text(row[1] if len(row) > 1 else None)
        if first.lower() in {"переменная", "variable", "name", "название"}:
            continue
        if first.lower() == "formula" and len(row) > 2:
            first = _cell_text(row[1])
            second = _cell_text(row[2])
        if first and second:
            second = second.lstrip("=").strip()
            formulas.append({"name": normalize_variable_name(first, f"formula_{len(formulas) + 1}"), "formula": second})
    return formulas


def _looks_like_new_test_format(row) -> bool:
    cells = [_cell_text(cell).lower() for cell in row[:6]]
    if any("свой" in cell or "комментар" in cell or "переменн" in cell for cell in cells):
        return True
    first = _cell_text(row[0] if len(row) > 0 else None)
    second = _cell_text(row[1] if len(row) > 1 else None)
    return first.isdigit() and bool(second)


def _looks_like_question_header(row) -> bool:
    cells = [_cell_text(cell).lower() for cell in row[:6]]
    return any(cell == "вопрос" for cell in cells) and any(cell.startswith("номер") for cell in cells)


def _build_question_schema(header) -> dict:
    option_columns = []
    value_columns = []
    button_columns = []
    for index, raw_cell in enumerate(header):
        if index < 6:
            continue
        cell = _cell_text(raw_cell).lower().replace("ё", "е")
        if not cell:
            continue
        if "значен" in cell or "балл" in cell or cell.startswith("value"):
            value_columns.append(index)
        elif "кнопк" in cell or "подпис" in cell or cell.startswith("button"):
            button_columns.append(index)
        elif "вариант" in cell or cell.startswith("option") or cell.startswith("answer"):
            option_columns.append(index)
    return {
        "option_columns": option_columns,
        "value_columns": value_columns,
        "button_columns": button_columns,
    }


def _decode_tabular_text(raw: bytes) -> str:
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            return raw.decode("cp1251")
        except UnicodeDecodeError as exc:
            raise ValueError("Не удалось определить кодировку файла. Сохраните CSV в UTF-8.") from exc


def _read_delimited_rows(content: str) -> list[list[str]]:
    sample = content[:8192]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters="|;,\t").delimiter
    except csv.Error:
        delimiter = max(("|", ";", ",", "\t"), key=lambda item: sample.count(item))
    return list(csv.reader(io.StringIO(content, newline=""), delimiter=delimiter))


def _read_xlsx_sheets(raw: bytes) -> list[tuple[str, list[tuple]]]:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

    with zipfile.ZipFile(io.BytesIO(raw)) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(f"{{{main_ns}}}si"):
                shared_strings.append("".join(node.text or "" for node in item.iter(f"{{{main_ns}}}t")))

        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships.findall(f"{{{package_rel_ns}}}Relationship")
        }

        result = []
        for sheet in workbook.findall(f".//{{{main_ns}}}sheet"):
            title = sheet.attrib.get("name", "Sheet")
            relation_id = sheet.attrib.get(f"{{{rel_ns}}}id")
            target = targets.get(relation_id, "")
            path = target.lstrip("/")
            if not path.startswith("xl/"):
                path = f"xl/{path}"
            sheet_root = ElementTree.fromstring(archive.read(path))
            rows = []
            for row in sheet_root.findall(f".//{{{main_ns}}}row"):
                values: list[object] = []
                for cell in row.findall(f"{{{main_ns}}}c"):
                    reference = cell.attrib.get("r", "A1")
                    letters = re.match(r"[A-Z]+", reference)
                    column_index = _xlsx_column_index(letters.group(0) if letters else "A")
                    while len(values) <= column_index:
                        values.append(None)
                    cell_type = cell.attrib.get("t")
                    formula = cell.find(f"{{{main_ns}}}f")
                    value_node = cell.find(f"{{{main_ns}}}v")
                    if formula is not None:
                        value: object = f"={formula.text or ''}"
                    elif cell_type == "inlineStr":
                        value = "".join(node.text or "" for node in cell.iter(f"{{{main_ns}}}t"))
                    elif value_node is None:
                        value = None
                    elif cell_type == "s":
                        value = shared_strings[int(value_node.text or "0")]
                    elif cell_type in {"str", "b"}:
                        value = value_node.text or ""
                    else:
                        raw_value = value_node.text or ""
                        try:
                            number = float(raw_value)
                            value = int(number) if number.is_integer() else number
                        except ValueError:
                            value = raw_value
                    values[column_index] = value
                rows.append(tuple(values))
            result.append((title, rows))
        return result


def _xlsx_column_index(letters: str) -> int:
    result = 0
    for char in letters:
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float_or_none(value):
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None
